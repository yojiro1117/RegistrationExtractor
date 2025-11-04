from __future__ import annotations

from pathlib import Path
from typing import List, Dict, Any, Tuple
import json
import os

from openpyxl import Workbook, load_workbook
from openpyxl.worksheet.worksheet import Worksheet


# ---------------------------
# 設定読込
# ---------------------------
def _read_settings() -> dict:
    settings_path = Path(__file__).resolve().parents[1] / "app" / "settings.json"
    default = {"output_sheet_name": "所有者一覧(最新_一括)"}
    try:
        with open(settings_path, "r", encoding="utf-8") as f:
            cfg = json.load(f)
            default.update(cfg or {})
    except Exception:
        pass
    return default


# ---------------------------
# ワークブック準備（空ファイル安全）
# ---------------------------
def _prepare_workbook(output_path: str | os.PathLike[str], sheet_name: str) -> Tuple[Workbook, Worksheet]:
    path = Path(output_path)

    if path.exists() and path.stat().st_size > 0:
        try:
            wb = load_workbook(path)
        except Exception:
            wb = Workbook()
    else:
        wb = Workbook()

    if sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
    else:
        if len(wb.sheetnames) == 1 and wb.active.max_row == 1 and wb.active.max_column == 1:
            ws = wb.active
            ws.title = sheet_name
        else:
            ws = wb.create_sheet(title=sheet_name)
    return wb, ws


# ---------------------------
# 固定ヘッダ（列順厳守）
# ---------------------------
# Complete list of all possible headers in the order they should appear in the
# generated Excel.  Do not modify this list without also updating the tests
# and the fields_profile.json default.  Individual users can choose a subset
# of these columns by editing app/fields_profile.json (see `_load_fields_profile`).
FIXED_HEADERS: List[str] = [
    "ファイル名","種別","事故簿フラグ","事故簿メモ",
    # 土地 表題部
    "所在","地番","地目","地積(㎡)原文","地積(㎡)数値","表題部_原因","表題部_原因日(原文)","表題部_原因日(YYYY-MM-DD)",
    # 建物 表題部（必要最低限）
    "家屋番号","種類","構造","床面積_1階(㎡)","表題部_原因_建物","表題部_原因日_建物(原文/規格化)",
    # 甲区（現所有者）
    "現権利者_持分(原文)","現権利者_持分(小数)","現権利者_氏名/名称","現権利者_住所",
    "現権利者_取得原因","現権利者_原因日(原文)","現権利者_原因日(YYYY-MM-DD)",
    "甲区_受付年月日","甲区_受付番号","備考_甲区",
    # 乙区（参考）
    "乙区_登記の目的","乙区_受付年月日","乙区_受付番号","乙区_権利者","乙区_原因","乙区_原因日(原文/規格化)","備考_乙区",
    # 法人（現行要約）
    "会社法人等番号","商号","本店","公告方法","会社成立年月日","目的(要約)","資本金",
    "発行可能株式総数","発行済株式数","機関（代表者/取締役等）","設置区分","最終登記日","備考_法人",
]


def _ensure_header(ws: Worksheet) -> None:
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(FIXED_HEADERS)


# ---------------------------
# 行生成（共有は行分割）
# ---------------------------
def _result_to_rows(result: Dict[str, Any]) -> List[List[Any]]:
    file_name = result.get("file_name", "")
    doc_type = result.get("type", "")

    accident_flag = bool(result.get("accident_flag", False))
    accident_memo = result.get("accident_memo", "")

    header = result.get("header", {}) or {}
    owners = result.get("owners", []) or []

    # 土地
    land_sozai = header.get("所在", "")
    land_chiban = header.get("地番", "")
    land_chimoku = header.get("地目", "")
    land_chiseki_raw = header.get("地積_原文", "")
    land_chiseki_num = header.get("地積_㎡", "")
    land_genin = header.get("表題部_原因", "")
    land_genin_day_raw = header.get("表題部_原因日_原文", "")
    land_genin_day_norm = header.get("表題部_原因日_規格化", "")

    # 建物（最低限）
    bld_kouban = header.get("家屋番号", "")
    bld_syurui = header.get("種類", "")
    bld_kouzou = header.get("構造", "")
    bld_menseki_1f = header.get("床面積_1階_㎡", "")
    bld_genin = header.get("表題部_原因_建物", "")
    bld_genin_day = header.get("表題部_原因日_建物", "")

    rows: List[List[Any]] = []
    if not owners:
        owners = [{}]  # 空でも1行

    for idx, o in enumerate(owners):
        # 共有2人目以降は共通列空欄
        common = [""] * len(FIXED_HEADERS)
        if idx == 0:
            common[:18] = [
                file_name,
                {"land": "土地", "building": "建物", "corporate": "法人"}.get(doc_type, ""),
                accident_flag,
                accident_memo,
                land_sozai, land_chiban, land_chimoku,
                land_chiseki_raw, land_chiseki_num,
                land_genin, land_genin_day_raw, land_genin_day_norm,
                bld_kouban, bld_syurui, bld_kouzou, bld_menseki_1f,
                bld_genin, bld_genin_day
            ]

        share_raw = o.get("share_raw", "")
        share_dec = o.get("share", "")
        name = o.get("name", "")
        address = o.get("address", "")
        acquire = o.get("acquire_reason", "")
        acquire_day_raw = o.get("acquire_day_raw", "")
        acquire_day_norm = o.get("acquire_day", "")
        kou_uketsuke_day = o.get("kou_uketsuke_day", "")
        kou_uketsuke_no = o.get("kou_uketsuke_no", "")
        kou_biko = o.get("kou_biko", "")

        # 乙区は参考（必要なら埋めてください）
        et_mokuteki = o.get("et_mokuteki", "")
        et_uketsuke_day = o.get("et_uketsuke_day", "")
        et_uketsuke_no = o.get("et_uketsuke_no", "")
        et_kenrisha = o.get("et_kenrisha", "")
        et_genin = o.get("et_genin", "")
        et_genin_day = o.get("et_genin_day", "")
        et_biko = o.get("et_biko", "")

        corp_no = header.get("会社法人等番号", "")
        corp_name = header.get("商号", "")
        corp_honten = header.get("本店", "")
        corp_koukoku = header.get("公告方法", "")
        corp_established = header.get("会社成立年月日", "")
        corp_purpose = header.get("目的(要約)", "")
        corp_capital = header.get("資本金", "")
        corp_total_iss = header.get("発行可能株式総数", "")
        corp_issued = header.get("発行済株式数", "")
        corp_org = header.get("機関", "")
        corp_setchi = header.get("設置区分", "")
        corp_final = header.get("最終登記日", "")
        corp_biko = header.get("備考_法人", "")

        row = [
            *common[:18],  # 共通部
            share_raw, share_dec,
            (f"{share_raw} " if share_raw else "") + name,  # 表示は「持分 原文 + 氏名」
            address, acquire, acquire_day_raw, acquire_day_norm,
            kou_uketsuke_day, kou_uketsuke_no, kou_biko,
            et_mokuteki, et_uketsuke_day, et_uketsuke_no, et_kenrisha, et_genin, et_genin_day, et_biko,
            corp_no, corp_name, corp_honten, corp_koukoku, corp_established, corp_purpose,
            corp_total_iss, corp_issued, corp_org, corp_setchi, corp_final, corp_biko
        ]
        # 列数調整
        if len(row) < len(FIXED_HEADERS):
            row += [""] * (len(FIXED_HEADERS) - len(row))
        elif len(row) > len(FIXED_HEADERS):
            row = row[:len(FIXED_HEADERS)]
        rows.append(row)

    return rows


# ---------------------------
# 公開関数
# ---------------------------
def _load_fields_profile(profile_path: str) -> Dict[str, bool]:
    """Load the column selection profile from a JSON file.

    The profile file should contain a mapping of header names (as defined in
    FIXED_HEADERS) to boolean values.  A value of True means the column will
    be included in the output Excel.  Missing keys default to True (i.e.,
    include all known columns unless explicitly turned off).  Extra keys are
    ignored.

    Parameters
    ----------
    profile_path : str
        Path to the JSON file with the column selection.

    Returns
    -------
    Dict[str, bool]
        A mapping of header -> bool indicating whether to include each header.
    """
    try:
        with open(profile_path, "r", encoding="utf-8") as f:
            data = json.load(f) or {}
        # ensure keys in our header set, default to True
        return {h: bool(data.get(h, True)) for h in FIXED_HEADERS}
    except Exception:
        # default: include all columns
        return {h: True for h in FIXED_HEADERS}


def write_results_to_excel(results: List[Dict[str, Any]], output_path: str, fields_profile_path: str = "app/fields_profile.json") -> None:
    """Write a list of extracted results to an Excel file.

    This function will respect the column selection specified in
    ``fields_profile_path``.  Only headers flagged as ``true`` will be
    written to the output Excel.  The order of columns follows
    ``FIXED_HEADERS``.

    Parameters
    ----------
    results : List[Dict[str, Any]]
        Parsed extraction results from the parsers.
    output_path : str
        Path to the output Excel file.
    fields_profile_path : str, optional
        Path to the JSON file specifying which columns to include.  Defaults
        to ``app/fields_profile.json``.
    """
    cfg = _read_settings()
    sheet_name = cfg.get("output_sheet_name", "所有者一覧(最新_一括)")

    # Load column selection
    profile = _load_fields_profile(fields_profile_path)
    # Compute the indices of columns to include
    active_indices = [i for i, h in enumerate(FIXED_HEADERS) if profile.get(h, True)]
    active_headers = [FIXED_HEADERS[i] for i in active_indices]

    wb, ws = _prepare_workbook(output_path, sheet_name)
    # Write header row only if the sheet is empty (A1 is None)
    if ws.max_row == 1 and ws.max_column == 1 and ws["A1"].value is None:
        ws.append(active_headers)
    # Append each result row filtered by active indices
    for result in results:
        full_rows = _result_to_rows(result)
        for row in full_rows:
            filtered = [row[i] for i in active_indices]
            ws.append(filtered)
    # Save workbook
    wb.save(output_path)
